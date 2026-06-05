from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional, List
from pydantic import BaseModel
import io, openpyxl
from openpyxl.styles import Font, PatternFill

from app.db.base import get_db
from app.models.finance import Account, JournalEntry, JournalLine, Expense, Budget
from app.models.user import User
from app.api.v1.endpoints.auth import get_current_user

router = APIRouter()

# ─── Accounts ─────────────────────────────────────────────────────────────────

class AccountCreate(BaseModel):
    account_code: str
    account_name: str
    account_type: str
    account_group: Optional[str] = None
    parent_id: Optional[int] = None
    description: Optional[str] = None
    is_group: bool = False
    opening_balance: float = 0

def acc_dict(a: Account):
    return {"id": a.id, "account_code": a.account_code, "account_name": a.account_name,
            "account_type": a.account_type, "account_group": a.account_group,
            "opening_balance": float(a.opening_balance or 0), "current_balance": float(a.current_balance or 0),
            "is_group": a.is_group, "is_active": a.is_active}

@router.get("/accounts")
async def list_accounts(account_type: Optional[str] = None, db: Session = Depends(get_db),
                         current_user: User = Depends(get_current_user)):
    q = db.query(Account).filter(Account.is_active == True)
    if account_type: q = q.filter(Account.account_type == account_type)
    items = q.order_by(Account.account_code).all()
    return [acc_dict(a) for a in items]

@router.post("/accounts")
async def create_account(data: AccountCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    a = Account(**data.dict()); db.add(a); db.commit(); db.refresh(a)
    return acc_dict(a)

@router.get("/accounts/export")
async def export_accounts(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    items = db.query(Account).filter(Account.is_active == True).order_by(Account.account_code).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Chart of Accounts"
    headers = ["Code", "Name", "Type", "Group", "Opening Balance", "Current Balance"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6366F1")
    for row, a in enumerate(items, 2):
        for col, val in enumerate([a.account_code, a.account_name, a.account_type, a.account_group,
                                    float(a.opening_balance or 0), float(a.current_balance or 0)], 1):
            ws.cell(row=row, column=col, value=val)
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=chart_of_accounts.xlsx"})

# ─── Journal Entries ──────────────────────────────────────────────────────────

class JournalLineIn(BaseModel):
    account_id: int
    description: Optional[str] = None
    debit_amount: float = 0
    credit_amount: float = 0

class JournalCreate(BaseModel):
    entry_date: str
    reference: Optional[str] = None
    description: Optional[str] = None
    entry_type: str = "manual"
    lines: List[JournalLineIn]

def je_dict(j: JournalEntry):
    return {"id": j.id, "entry_number": j.entry_number, "entry_date": j.entry_date.isoformat() if j.entry_date else None,
            "reference": j.reference, "description": j.description, "status": j.status,
            "total_debit": float(j.total_debit or 0), "total_credit": float(j.total_credit or 0),
            "created_at": j.created_at.isoformat() if j.created_at else None}

@router.get("/journals")
async def list_journals(skip: int = 0, limit: int = 50, status: Optional[str] = None,
                         db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    q = db.query(JournalEntry).filter(JournalEntry.is_active == True)
    if status: q = q.filter(JournalEntry.status == status)
    total = q.count()
    items = q.order_by(JournalEntry.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [je_dict(j) for j in items]}

@router.post("/journals")
async def create_journal(data: JournalCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from datetime import datetime
    count = db.query(func.count(JournalEntry.id)).scalar()
    total_debit = sum(l.debit_amount for l in data.lines)
    total_credit = sum(l.credit_amount for l in data.lines)
    if abs(total_debit - total_credit) > 0.01:
        raise HTTPException(400, "Debit and Credit must balance")
    je = JournalEntry(
        entry_number=f"JE-{(count or 0) + 1:05d}",
        entry_date=datetime.strptime(data.entry_date, "%Y-%m-%d"),
        reference=data.reference, description=data.description,
        entry_type=data.entry_type, total_debit=total_debit, total_credit=total_credit,
        created_by=current_user.id
    )
    for line in data.lines:
        jl = JournalLine(account_id=line.account_id, description=line.description,
                          debit_amount=line.debit_amount, credit_amount=line.credit_amount)
        je.lines.append(jl)
    db.add(je); db.commit(); db.refresh(je)
    return je_dict(je)

@router.get("/journals/export")
async def export_journals(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    items = db.query(JournalEntry).filter(JournalEntry.is_active == True).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Journal Entries"
    headers = ["Entry #", "Date", "Reference", "Description", "Debit", "Credit", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6366F1")
    for row, j in enumerate(items, 2):
        for col, val in enumerate([j.entry_number, j.entry_date.strftime("%Y-%m-%d") if j.entry_date else "",
                                    j.reference, j.description, float(j.total_debit or 0),
                                    float(j.total_credit or 0), j.status], 1):
            ws.cell(row=row, column=col, value=val)
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=journal_entries.xlsx"})

@router.put("/journals/{je_id}/post")
async def post_journal(je_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    j = db.query(JournalEntry).filter(JournalEntry.id == je_id).first()
    if not j: raise HTTPException(404, "Not found")
    if j.status == "posted": raise HTTPException(400, "Already posted")
    j.status = "posted"; j.approved_by = current_user.id
    for line in j.lines:
        acc = db.query(Account).filter(Account.id == line.account_id).first()
        if acc:
            current = float(acc.current_balance or 0)
            if acc.account_type in ["asset", "expense"]:
                acc.current_balance = current + float(line.debit_amount) - float(line.credit_amount)
            else:
                acc.current_balance = current + float(line.credit_amount) - float(line.debit_amount)
    db.commit(); return {"message": "Posted"}

# ─── Expenses ─────────────────────────────────────────────────────────────────

class ExpenseCreate(BaseModel):
    expense_date: str
    category: str
    description: Optional[str] = None
    amount: float
    currency: str = "INR"
    payment_mode: Optional[str] = None
    vendor_name: Optional[str] = None
    employee_id: Optional[int] = None
    project_id: Optional[int] = None

@router.get("/expenses")
async def list_expenses(skip: int = 0, limit: int = 50, status: Optional[str] = None,
                         db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    q = db.query(Expense)
    if status: q = q.filter(Expense.status == status)
    total = q.count()
    items = q.order_by(Expense.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [{"id": e.id, "expense_number": e.expense_number,
            "expense_date": e.expense_date.isoformat() if e.expense_date else None,
            "category": e.category, "description": e.description, "amount": float(e.amount),
            "status": e.status, "payment_mode": e.payment_mode} for e in items]}

@router.post("/expenses")
async def create_expense(data: ExpenseCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from datetime import datetime
    count = db.query(func.count(Expense.id)).scalar()
    exp = Expense(expense_number=f"EXP-{(count or 0) + 1:05d}",
                   expense_date=datetime.strptime(data.expense_date, "%Y-%m-%d"),
                   category=data.category, description=data.description, amount=data.amount,
                   total_amount=data.amount, currency=data.currency, payment_mode=data.payment_mode,
                   vendor_name=data.vendor_name, employee_id=data.employee_id, project_id=data.project_id)
    db.add(exp); db.commit(); db.refresh(exp)
    return {"id": exp.id, "expense_number": exp.expense_number}

@router.get("/expenses/export")
async def export_expenses(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    items = db.query(Expense).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Expenses"
    headers = ["Expense #", "Date", "Category", "Description", "Amount", "Vendor", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6366F1")
    for row, e in enumerate(items, 2):
        for col, val in enumerate([e.expense_number, e.expense_date.strftime("%Y-%m-%d") if e.expense_date else "",
                                    e.category, e.description, float(e.amount), e.vendor_name, e.status], 1):
            ws.cell(row=row, column=col, value=val)
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=expenses.xlsx"})

# ─── Reports ─────────────────────────────────────────────────────────────────

@router.get("/reports/trial-balance")
async def trial_balance(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    accounts = db.query(Account).filter(Account.is_active == True, Account.is_group == False).all()
    return [{"code": a.account_code, "name": a.account_name, "type": a.account_type,
             "opening": float(a.opening_balance or 0), "current": float(a.current_balance or 0),
             "debit": float(a.current_balance) if a.account_type in ["asset","expense"] and a.current_balance > 0 else 0,
             "credit": float(abs(a.current_balance)) if a.account_type not in ["asset","expense"] else 0} for a in accounts]

@router.get("/reports/pl")
async def profit_loss(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    revenue = db.query(func.sum(Account.current_balance)).filter(Account.account_type == "revenue").scalar() or 0
    expense = db.query(func.sum(Account.current_balance)).filter(Account.account_type == "expense").scalar() or 0
    return {"revenue": float(revenue), "expense": float(expense), "net_profit": float(revenue) - float(expense)}
