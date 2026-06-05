from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional, List
from pydantic import BaseModel
import io, openpyxl
from openpyxl.styles import Font, PatternFill

from app.db.base import get_db
from app.models.project import Project, ProjectTask, Milestone, Timesheet
from app.models.user import User
from app.api.v1.endpoints.auth import get_current_user

router = APIRouter()

# ── Schemas ────────────────────────────────────────────────────────────────────

class ProjectCreate(BaseModel):
    name: str
    description: Optional[str] = None
    project_type: str = "it_services"
    customer_id: Optional[int] = None
    status: str = "planning"
    priority: str = "medium"
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    budget: float = 0
    project_manager_id: Optional[int] = None

class TaskCreate(BaseModel):
    project_id: int
    title: str
    description: Optional[str] = None
    task_type: str = "task"
    status: str = "todo"
    priority: str = "medium"
    assigned_to: Optional[int] = None
    due_date: Optional[str] = None
    estimated_hours: Optional[float] = None

class TimesheetCreate(BaseModel):
    project_id: int
    task_id: Optional[int] = None
    date: str
    hours: float
    billable: bool = True
    description: Optional[str] = None

def proj_dict(p: Project):
    return {"id": p.id, "project_number": p.project_number, "name": p.name,
            "description": p.description, "project_type": p.project_type,
            "customer_id": p.customer_id, "status": p.status, "priority": p.priority,
            "start_date": str(p.start_date) if p.start_date else None,
            "end_date": str(p.end_date) if p.end_date else None,
            "budget": float(p.budget or 0), "actual_cost": float(p.actual_cost or 0),
            "progress_percent": p.progress_percent, "project_manager_id": p.project_manager_id,
            "is_active": p.is_active, "created_at": p.created_at.isoformat() if p.created_at else None}

# ── Projects — static routes FIRST (before /{project_id}) ─────────────────────

@router.get("/")
async def list_projects(skip: int = 0, limit: int = 50, status: Optional[str] = None,
                         project_type: Optional[str] = None, db: Session = Depends(get_db),
                         current_user: User = Depends(get_current_user)):
    q = db.query(Project).filter(Project.is_active == True)
    if status: q = q.filter(Project.status == status)
    if project_type: q = q.filter(Project.project_type == project_type)
    total = q.count()
    items = q.order_by(Project.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [proj_dict(p) for p in items]}

@router.post("/")
async def create_project(data: ProjectCreate, db: Session = Depends(get_db),
                          current_user: User = Depends(get_current_user)):
    from datetime import datetime as dt
    count = db.query(func.count(Project.id)).scalar()
    d = data.dict()
    if d.get("start_date"): d["start_date"] = dt.strptime(d["start_date"], "%Y-%m-%d").date()
    if d.get("end_date"): d["end_date"] = dt.strptime(d["end_date"], "%Y-%m-%d").date()
    proj = Project(project_number=f"PRJ-{(count or 0) + 1:05d}", **d)
    db.add(proj); db.commit(); db.refresh(proj)
    return proj_dict(proj)

@router.get("/stats")
async def project_stats(db: Session = Depends(get_db),
                         current_user: User = Depends(get_current_user)):
    total = db.query(func.count(Project.id)).filter(Project.is_active == True).scalar()
    by_status = dict(db.query(Project.status, func.count(Project.id)).filter(Project.is_active == True).group_by(Project.status).all())
    total_budget = db.query(func.sum(Project.budget)).filter(Project.is_active == True).scalar()
    return {"total": total, "by_status": by_status, "total_budget": float(total_budget or 0)}

@router.get("/export")
async def export_projects(db: Session = Depends(get_db),
                           current_user: User = Depends(get_current_user)):
    items = db.query(Project).filter(Project.is_active == True).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Projects"
    headers = ["Project #", "Name", "Type", "Status", "Priority", "Start", "End", "Budget"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6366F1")
    for row, p in enumerate(items, 2):
        for col, val in enumerate([p.project_number, p.name, p.project_type, p.status, p.priority,
                                    str(p.start_date) if p.start_date else "",
                                    str(p.end_date) if p.end_date else "",
                                    float(p.budget or 0)], 1):
            ws.cell(row=row, column=col, value=val)
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=projects.xlsx"})

# ── Tasks — static routes BEFORE /{project_id} ────────────────────────────────

@router.get("/tasks/all")
async def list_all_tasks(skip: int = 0, limit: int = 50, project_id: Optional[int] = None,
                          status: Optional[str] = None, assigned_to: Optional[int] = None,
                          db: Session = Depends(get_db),
                          current_user: User = Depends(get_current_user)):
    q = db.query(ProjectTask).filter(ProjectTask.is_active == True)
    if project_id: q = q.filter(ProjectTask.project_id == project_id)
    if status: q = q.filter(ProjectTask.status == status)
    if assigned_to: q = q.filter(ProjectTask.assigned_to == assigned_to)
    total = q.count()
    items = q.order_by(ProjectTask.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [{"id": t.id, "project_id": t.project_id, "title": t.title,
            "task_type": t.task_type, "status": t.status, "priority": t.priority,
            "assigned_to": t.assigned_to, "progress_percent": t.progress_percent,
            "estimated_hours": t.estimated_hours, "actual_hours": t.actual_hours} for t in items]}

@router.post("/tasks")
async def create_task(data: TaskCreate, db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user)):
    from datetime import datetime
    t = ProjectTask(**{k: v for k, v in data.dict().items() if k != "due_date"})
    if data.due_date:
        t.due_date = datetime.strptime(data.due_date, "%Y-%m-%d").date()
    db.add(t); db.commit(); db.refresh(t)
    return {"id": t.id, "title": t.title, "status": t.status}

@router.put("/tasks/{task_id}")
async def update_task(task_id: int, data: dict, db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user)):
    t = db.query(ProjectTask).filter(ProjectTask.id == task_id).first()
    if not t: raise HTTPException(404, "Not found")
    for k, v in data.items():
        if hasattr(t, k): setattr(t, k, v)
    db.commit(); return {"id": t.id, "status": t.status}

# ── Timesheets — static routes BEFORE /{project_id} ───────────────────────────

@router.get("/timesheets")
async def list_timesheets(project_id: Optional[int] = None, skip: int = 0, limit: int = 200,
                           db: Session = Depends(get_db),
                           current_user: User = Depends(get_current_user)):
    q = db.query(Timesheet)
    if project_id: q = q.filter(Timesheet.project_id == project_id)
    total = q.count()
    items = q.order_by(Timesheet.date.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [{"id": t.id, "project_id": t.project_id,
            "task_id": t.task_id, "date": str(t.date), "hours": float(t.hours or 0),
            "billable": t.billable, "description": t.description,
            "user_id": t.user_id} for t in items]}

@router.post("/timesheets")
async def log_timesheet(data: TimesheetCreate, db: Session = Depends(get_db),
                         current_user: User = Depends(get_current_user)):
    from datetime import datetime
    t = Timesheet(project_id=data.project_id, task_id=data.task_id,
                  user_id=current_user.id, hours=data.hours, billable=data.billable,
                  description=data.description,
                  date=datetime.strptime(data.date, "%Y-%m-%d").date())
    db.add(t); db.commit(); db.refresh(t)
    return {"id": t.id, "hours": float(t.hours), "date": str(t.date)}

# ── Projects — parameterised routes LAST ──────────────────────────────────────

@router.get("/{project_id}")
async def get_project(project_id: int, db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user)):
    p = db.query(Project).filter(Project.id == project_id).first()
    if not p: raise HTTPException(404, "Not found")
    d = proj_dict(p)
    d["tasks"] = [{"id": t.id, "title": t.title, "status": t.status, "priority": t.priority,
                   "assigned_to": t.assigned_to, "due_date": str(t.due_date) if t.due_date else None,
                   "progress_percent": t.progress_percent} for t in p.tasks]
    d["milestones"] = [{"id": m.id, "name": m.name, "due_date": str(m.due_date) if m.due_date else None,
                        "status": m.status} for m in p.milestones]
    return d

@router.put("/{project_id}")
async def update_project(project_id: int, data: dict, db: Session = Depends(get_db),
                          current_user: User = Depends(get_current_user)):
    p = db.query(Project).filter(Project.id == project_id).first()
    if not p: raise HTTPException(404, "Not found")
    for k, v in data.items():
        if hasattr(p, k): setattr(p, k, v)
    db.commit(); return proj_dict(p)

@router.delete("/{project_id}")
async def delete_project(project_id: int, db: Session = Depends(get_db),
                          current_user: User = Depends(get_current_user)):
    p = db.query(Project).filter(Project.id == project_id).first()
    if not p: raise HTTPException(404, "Not found")
    p.is_active = False; db.commit(); return {"message": "Deleted"}
