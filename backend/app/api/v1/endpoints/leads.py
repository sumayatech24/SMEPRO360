from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from pydantic import BaseModel
from datetime import datetime
import io

from app.db.base import get_db
from app.models.lead import Lead, LeadActivity, Campaign
from app.models.user import User
from app.api.v1.endpoints.auth import get_current_user

router = APIRouter()

class LeadCreate(BaseModel):
    first_name: str
    last_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    company: Optional[str] = None
    designation: Optional[str] = None
    industry: Optional[str] = None
    source: Optional[str] = None
    status: str = "new"
    priority: str = "medium"
    annual_revenue: Optional[float] = None
    city: Optional[str] = None
    state: Optional[str] = None
    description: Optional[str] = None
    assigned_to: Optional[int] = None
    campaign_id: Optional[int] = None

class LeadUpdate(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    company: Optional[str] = None
    status: Optional[str] = None
    priority: Optional[str] = None
    assigned_to: Optional[int] = None
    description: Optional[str] = None

def lead_to_dict(lead: Lead):
    return {
        "id": lead.id,
        "lead_number": lead.lead_number,
        "first_name": lead.first_name,
        "last_name": lead.last_name,
        "email": lead.email,
        "phone": lead.phone,
        "company": lead.company,
        "designation": lead.designation,
        "industry": lead.industry,
        "source": lead.source,
        "status": lead.status,
        "priority": lead.priority,
        "annual_revenue": float(lead.annual_revenue) if lead.annual_revenue else None,
        "city": lead.city,
        "state": lead.state,
        "assigned_to": lead.assigned_to,
        "campaign_id": lead.campaign_id,
        "description": lead.description,
        "is_active": lead.is_active,
        "created_at": lead.created_at.isoformat() if lead.created_at else None,
        "updated_at": lead.updated_at.isoformat() if lead.updated_at else None,
    }

@router.get("/")
async def list_leads(
    skip: int = 0, limit: int = 50,
    search: Optional[str] = None,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    assigned_to: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(Lead).filter(Lead.is_active == True)
    if search:
        query = query.filter(
            Lead.first_name.ilike(f"%{search}%") |
            Lead.last_name.ilike(f"%{search}%") |
            Lead.company.ilike(f"%{search}%") |
            Lead.email.ilike(f"%{search}%")
        )
    if status:
        query = query.filter(Lead.status == status)
    if priority:
        query = query.filter(Lead.priority == priority)
    if assigned_to:
        query = query.filter(Lead.assigned_to == assigned_to)
    total = query.count()
    leads = query.order_by(Lead.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [lead_to_dict(l) for l in leads]}

@router.post("/")
async def create_lead(data: LeadCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    # Generate lead number
    count = db.query(func.count(Lead.id)).scalar()
    lead_number = f"LD-{(count or 0) + 1:05d}"
    lead = Lead(lead_number=lead_number, **data.dict())
    db.add(lead)
    db.commit()
    db.refresh(lead)
    return lead_to_dict(lead)

@router.get("/stats")
async def lead_stats(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    total = db.query(func.count(Lead.id)).filter(Lead.is_active == True).scalar()
    by_status = db.query(Lead.status, func.count(Lead.id)).filter(Lead.is_active == True).group_by(Lead.status).all()
    by_source = db.query(Lead.source, func.count(Lead.id)).filter(Lead.is_active == True).group_by(Lead.source).all()
    return {
        "total": total,
        "by_status": {s: c for s, c in by_status if s},
        "by_source": {s: c for s, c in by_source if s},
    }

@router.get("/export")
async def export_leads(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    leads = db.query(Lead).filter(Lead.is_active == True).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    headers = ["Lead #", "First Name", "Last Name", "Email", "Phone", "Company", "Industry", "Source", "Status", "Priority", "City", "State", "Created"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6366F1")
        cell.alignment = Alignment(horizontal="center")
    for row, lead in enumerate(leads, 2):
        ws.cell(row=row, column=1, value=lead.lead_number)
        ws.cell(row=row, column=2, value=lead.first_name)
        ws.cell(row=row, column=3, value=lead.last_name)
        ws.cell(row=row, column=4, value=lead.email)
        ws.cell(row=row, column=5, value=lead.phone)
        ws.cell(row=row, column=6, value=lead.company)
        ws.cell(row=row, column=7, value=lead.industry)
        ws.cell(row=row, column=8, value=lead.source)
        ws.cell(row=row, column=9, value=lead.status)
        ws.cell(row=row, column=10, value=lead.priority)
        ws.cell(row=row, column=11, value=lead.city)
        ws.cell(row=row, column=12, value=lead.state)
        ws.cell(row=row, column=13, value=lead.created_at.strftime("%Y-%m-%d") if lead.created_at else "")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=leads.xlsx"})

@router.get("/{lead_id}")
async def get_lead(lead_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    return lead_to_dict(lead)

@router.put("/{lead_id}")
async def update_lead(lead_id: int, data: LeadUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    for k, v in data.dict(exclude_none=True).items():
        setattr(lead, k, v)
    db.commit()
    db.refresh(lead)
    return lead_to_dict(lead)

@router.delete("/{lead_id}")
async def delete_lead(lead_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    lead.is_active = False
    db.commit()
    return {"message": "Lead deleted"}

@router.post("/{lead_id}/convert")
async def convert_lead(lead_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    from app.models.crm import Customer
    from sqlalchemy import func as sqlfunc
    count = db.query(sqlfunc.count(Customer.id)).scalar()
    customer = Customer(
        customer_number=f"CUS-{(count or 0) + 1:05d}",
        company_name=lead.company or f"{lead.first_name} {lead.last_name or ''}".strip(),
        industry=lead.industry,
        city=lead.city,
        state=lead.state,
        phone=lead.phone,
        email=lead.email,
        source=lead.source,
        lead_id=lead.id,
    )
    db.add(customer)
    lead.status = "converted"
    lead.converted_at = datetime.utcnow()
    db.commit()
    return {"message": "Lead converted to customer", "customer_id": customer.id}

# Campaigns
@router.get("/campaigns/list")
async def list_campaigns(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    campaigns = db.query(Campaign).filter(Campaign.is_active == True).all()
    return [{"id": c.id, "name": c.name, "campaign_type": c.campaign_type, "status": c.status,
             "budget": float(c.budget or 0), "target_leads": c.target_leads} for c in campaigns]
