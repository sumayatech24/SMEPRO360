from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional, List
from pydantic import BaseModel
import io, openpyxl
from openpyxl.styles import Font, PatternFill

from app.db.base import get_db
from app.models.procurement import Vendor, PurchaseOrder, PurchaseOrderItem, GoodsReceipt
from app.models.user import User
from app.api.v1.endpoints.auth import get_current_user

router = APIRouter()

class VendorCreate(BaseModel):
    company_name: str
    vendor_type: str = "supplier"
    gstin: Optional[str] = None
    pan: Optional[str] = None
    msme_number: Optional[str] = None
    msme_category: Optional[str] = None
    credit_days: int = 30
    city: Optional[str] = None
    state: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    bank_name: Optional[str] = None
    bank_account: Optional[str] = None
    bank_ifsc: Optional[str] = None

def vendor_dict(v: Vendor):
    return {"id": v.id, "vendor_number": v.vendor_number, "company_name": v.company_name,
            "vendor_type": v.vendor_type, "gstin": v.gstin, "pan": v.pan,
            "credit_days": v.credit_days, "city": v.city, "state": v.state,
            "phone": v.phone, "email": v.email, "rating": v.rating, "status": v.status,
            "msme_category": v.msme_category, "is_active": v.is_active,
            "created_at": v.created_at.isoformat() if v.created_at else None}

@router.get("/vendors")
async def list_vendors(skip: int = 0, limit: int = 50, search: Optional[str] = None,
                        db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    q = db.query(Vendor).filter(Vendor.is_active == True)
    if search: q = q.filter(Vendor.company_name.ilike(f"%{search}%"))
    total = q.count()
    items = q.order_by(Vendor.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [vendor_dict(v) for v in items]}

@router.post("/vendors")
async def create_vendor(data: VendorCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    count = db.query(func.count(Vendor.id)).scalar()
    v = Vendor(vendor_number=f"VEN-{(count or 0) + 1:05d}", **data.dict())
    db.add(v); db.commit(); db.refresh(v)
    return vendor_dict(v)

@router.get("/vendors/export")
async def export_vendors(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    items = db.query(Vendor).filter(Vendor.is_active == True).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Vendors"
    headers = ["Vendor #", "Company", "Type", "GSTIN", "MSME Cat", "City", "State", "Phone", "Email", "Rating"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6366F1")
    for row, v in enumerate(items, 2):
        for col, val in enumerate([v.vendor_number, v.company_name, v.vendor_type, v.gstin,
                                    v.msme_category, v.city, v.state, v.phone, v.email, v.rating], 1):
            ws.cell(row=row, column=col, value=val)
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=vendors.xlsx"})

@router.get("/vendors/{vendor_id}")
async def get_vendor(vendor_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    v = db.query(Vendor).filter(Vendor.id == vendor_id).first()
    if not v: raise HTTPException(404, "Not found")
    return vendor_dict(v)

@router.put("/vendors/{vendor_id}")
async def update_vendor(vendor_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    v = db.query(Vendor).filter(Vendor.id == vendor_id).first()
    if not v: raise HTTPException(404, "Not found")
    for k, val in data.items():
        if hasattr(v, k): setattr(v, k, val)
    db.commit(); return vendor_dict(v)

@router.delete("/vendors/{vendor_id}")
async def delete_vendor(vendor_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    v = db.query(Vendor).filter(Vendor.id == vendor_id).first()
    if not v: raise HTTPException(404, "Not found")
    v.is_active = False; db.commit(); return {"message": "Deleted"}

# ─── Purchase Orders ─────────────────────────────────────────────────────────

class POItemIn(BaseModel):
    product_id: Optional[int] = None
    description: str
    quantity: float
    unit: Optional[str] = "nos"
    unit_price: float
    tax_percent: float = 18

class POCreate(BaseModel):
    vendor_id: int
    expected_delivery: Optional[str] = None
    delivery_address: Optional[str] = None
    payment_terms: Optional[str] = None
    notes: Optional[str] = None
    items: List[POItemIn] = []

def po_dict(o: PurchaseOrder):
    return {"id": o.id, "po_number": o.po_number, "vendor_id": o.vendor_id, "status": o.status,
            "order_date": o.order_date.isoformat() if o.order_date else None,
            "total_amount": float(o.total_amount or 0), "payment_terms": o.payment_terms,
            "created_at": o.created_at.isoformat() if o.created_at else None}

@router.get("/orders")
async def list_po(skip: int = 0, limit: int = 50, status: Optional[str] = None,
                   vendor_id: Optional[int] = None, db: Session = Depends(get_db),
                   current_user: User = Depends(get_current_user)):
    q = db.query(PurchaseOrder).filter(PurchaseOrder.is_active == True)
    if status: q = q.filter(PurchaseOrder.status == status)
    if vendor_id: q = q.filter(PurchaseOrder.vendor_id == vendor_id)
    total = q.count()
    items = q.order_by(PurchaseOrder.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [po_dict(o) for o in items]}

@router.post("/orders")
async def create_po(data: POCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    count = db.query(func.count(PurchaseOrder.id)).scalar()
    po = PurchaseOrder(po_number=f"PO-{(count or 0) + 1:05d}", vendor_id=data.vendor_id,
                        payment_terms=data.payment_terms, notes=data.notes, created_by=current_user.id)
    subtotal = tax_total = 0
    for item_data in data.items:
        tax = item_data.quantity * item_data.unit_price * item_data.tax_percent / 100
        line = item_data.quantity * item_data.unit_price
        item = PurchaseOrderItem(product_id=item_data.product_id, description=item_data.description,
                                  quantity=item_data.quantity, unit=item_data.unit,
                                  unit_price=item_data.unit_price, tax_percent=item_data.tax_percent,
                                  tax_amount=tax, line_total=line + tax)
        po.items.append(item)
        subtotal += line; tax_total += tax
    po.subtotal = subtotal; po.tax_amount = tax_total; po.total_amount = subtotal + tax_total
    db.add(po); db.commit(); db.refresh(po)
    return po_dict(po)

@router.get("/orders/export")
async def export_po(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    items = db.query(PurchaseOrder).filter(PurchaseOrder.is_active == True).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Purchase Orders"
    headers = ["PO #", "Vendor ID", "Status", "Order Date", "Total Amount"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6366F1")
    for row, o in enumerate(items, 2):
        for col, val in enumerate([o.po_number, o.vendor_id, o.status,
                                    o.order_date.strftime("%Y-%m-%d") if o.order_date else "",
                                    float(o.total_amount or 0)], 1):
            ws.cell(row=row, column=col, value=val)
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=purchase_orders.xlsx"})

@router.get("/orders/{po_id}")
async def get_po(po_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    o = db.query(PurchaseOrder).filter(PurchaseOrder.id == po_id).first()
    if not o: raise HTTPException(404, "Not found")
    d = po_dict(o)
    d["items"] = [{"id": i.id, "description": i.description, "quantity": i.quantity,
                   "unit_price": float(i.unit_price), "line_total": float(i.line_total or 0)} for i in o.items]
    return d

@router.put("/orders/{po_id}")
async def update_po(po_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    o = db.query(PurchaseOrder).filter(PurchaseOrder.id == po_id).first()
    if not o: raise HTTPException(404, "Not found")
    for k, v in data.items():
        if hasattr(o, k): setattr(o, k, v)
    db.commit(); return po_dict(o)

@router.delete("/orders/{po_id}")
async def delete_po(po_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    o = db.query(PurchaseOrder).filter(PurchaseOrder.id == po_id).first()
    if not o: raise HTTPException(404, "Not found")
    o.is_active = False; db.commit(); return {"message": "Deleted"}
