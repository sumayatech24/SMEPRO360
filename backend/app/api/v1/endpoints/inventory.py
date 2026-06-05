from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from pydantic import BaseModel
import io, openpyxl
from openpyxl.styles import Font, PatternFill

from app.db.base import get_db
from app.models.inventory import Product, Category, Warehouse, StockLevel, StockMovement
from app.models.user import User
from app.api.v1.endpoints.auth import get_current_user

router = APIRouter()

class ProductCreate(BaseModel):
    sku: str
    name: str
    description: Optional[str] = None
    category_id: Optional[int] = None
    product_type: str = "finished"
    unit: str = "nos"
    hsn_code: Optional[str] = None
    cost_price: float = 0
    selling_price: float = 0
    mrp: float = 0
    tax_percent: float = 18
    reorder_level: float = 0
    reorder_quantity: float = 0

def prod_dict(p: Product):
    return {"id": p.id, "sku": p.sku, "name": p.name, "description": p.description,
            "category_id": p.category_id, "product_type": p.product_type, "unit": p.unit,
            "hsn_code": p.hsn_code, "cost_price": float(p.cost_price or 0),
            "selling_price": float(p.selling_price or 0), "mrp": float(p.mrp or 0),
            "tax_percent": p.tax_percent, "reorder_level": p.reorder_level, "is_active": p.is_active,
            "created_at": p.created_at.isoformat() if p.created_at else None}

@router.get("/products")
async def list_products(skip: int = 0, limit: int = 50, search: Optional[str] = None,
                         category_id: Optional[int] = None, product_type: Optional[str] = None,
                         db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    q = db.query(Product).filter(Product.is_active == True)
    if search: q = q.filter(Product.name.ilike(f"%{search}%") | Product.sku.ilike(f"%{search}%"))
    if category_id: q = q.filter(Product.category_id == category_id)
    if product_type: q = q.filter(Product.product_type == product_type)
    total = q.count()
    items = q.order_by(Product.name).offset(skip).limit(limit).all()
    return {"total": total, "items": [prod_dict(p) for p in items]}

@router.post("/products")
async def create_product(data: ProductCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if db.query(Product).filter(Product.sku == data.sku).first():
        raise HTTPException(400, "SKU already exists")
    p = Product(**data.dict()); db.add(p); db.commit(); db.refresh(p)
    return prod_dict(p)

@router.get("/products/export")
async def export_products(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    items = db.query(Product).filter(Product.is_active == True).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Products"
    headers = ["SKU", "Name", "Type", "Unit", "HSN", "Cost Price", "Selling Price", "MRP", "Tax %", "Reorder Level"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6366F1")
    for row, p in enumerate(items, 2):
        for col, val in enumerate([p.sku, p.name, p.product_type, p.unit, p.hsn_code,
                                    float(p.cost_price or 0), float(p.selling_price or 0),
                                    float(p.mrp or 0), p.tax_percent, p.reorder_level], 1):
            ws.cell(row=row, column=col, value=val)
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=products.xlsx"})

@router.get("/products/{product_id}")
async def get_product(product_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    p = db.query(Product).filter(Product.id == product_id).first()
    if not p: raise HTTPException(404, "Not found")
    return prod_dict(p)

@router.put("/products/{product_id}")
async def update_product(product_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    p = db.query(Product).filter(Product.id == product_id).first()
    if not p: raise HTTPException(404, "Not found")
    for k, v in data.items():
        if hasattr(p, k): setattr(p, k, v)
    db.commit(); return prod_dict(p)

@router.delete("/products/{product_id}")
async def delete_product(product_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    p = db.query(Product).filter(Product.id == product_id).first()
    if not p: raise HTTPException(404, "Not found")
    p.is_active = False; db.commit(); return {"message": "Deleted"}

# ─── Categories ───────────────────────────────────────────────────────────────

@router.get("/categories")
async def list_categories(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    cats = db.query(Category).filter(Category.is_active == True).all()
    return [{"id": c.id, "name": c.name, "code": c.code, "parent_id": c.parent_id} for c in cats]

@router.post("/categories")
async def create_category(data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    c = Category(**{k: v for k, v in data.items() if hasattr(Category, k)})
    db.add(c); db.commit(); db.refresh(c)
    return {"id": c.id, "name": c.name}

# ─── Warehouses ───────────────────────────────────────────────────────────────

@router.get("/warehouses")
async def list_warehouses(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    whs = db.query(Warehouse).filter(Warehouse.is_active == True).all()
    return [{"id": w.id, "name": w.name, "code": w.code, "city": w.city, "state": w.state} for w in whs]

@router.post("/warehouses")
async def create_warehouse(data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    w = Warehouse(**{k: v for k, v in data.items() if hasattr(Warehouse, k)})
    db.add(w); db.commit(); db.refresh(w)
    return {"id": w.id, "name": w.name}

# ─── Stock ────────────────────────────────────────────────────────────────────

@router.get("/stock")
async def get_stock(warehouse_id: Optional[int] = None, product_id: Optional[int] = None,
                     skip: int = 0, limit: int = 100, db: Session = Depends(get_db),
                     current_user: User = Depends(get_current_user)):
    q = db.query(StockLevel)
    if warehouse_id: q = q.filter(StockLevel.warehouse_id == warehouse_id)
    if product_id: q = q.filter(StockLevel.product_id == product_id)
    total = q.count()
    items = q.offset(skip).limit(limit).all()
    result = []
    for s in items:
        product = db.query(Product).filter(Product.id == s.product_id).first()
        warehouse = db.query(Warehouse).filter(Warehouse.id == s.warehouse_id).first()
        result.append({"id": s.id, "product_id": s.product_id,
                       "product_name": product.name if product else "", "product_sku": product.sku if product else "",
                       "warehouse_id": s.warehouse_id, "warehouse_name": warehouse.name if warehouse else "",
                       "quantity_on_hand": s.quantity_on_hand, "quantity_available": s.quantity_available,
                       "quantity_reserved": s.quantity_reserved, "reorder_level": product.reorder_level if product else 0,
                       "below_reorder": s.quantity_on_hand < (product.reorder_level if product else 0)})
    return {"total": total, "items": result}

@router.get("/stock/export")
async def export_stock(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    items = db.query(StockLevel).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Stock"
    headers = ["Product ID", "SKU", "Product Name", "Warehouse", "On Hand", "Available", "Reserved"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6366F1")
    for row, s in enumerate(items, 2):
        product = db.query(Product).filter(Product.id == s.product_id).first()
        warehouse = db.query(Warehouse).filter(Warehouse.id == s.warehouse_id).first()
        for col, val in enumerate([s.product_id, product.sku if product else "", product.name if product else "",
                                    warehouse.name if warehouse else "",
                                    s.quantity_on_hand, s.quantity_available, s.quantity_reserved], 1):
            ws.cell(row=row, column=col, value=val)
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=stock.xlsx"})

@router.post("/stock/adjust")
async def adjust_stock(data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    product_id = data.get("product_id")
    warehouse_id = data.get("warehouse_id")
    quantity = data.get("quantity", 0)
    movement_type = data.get("movement_type", "adjustment")
    stock = db.query(StockLevel).filter(StockLevel.product_id == product_id, StockLevel.warehouse_id == warehouse_id).first()
    if not stock:
        stock = StockLevel(product_id=product_id, warehouse_id=warehouse_id, quantity_on_hand=0, quantity_available=0, quantity_reserved=0)
        db.add(stock)
    if movement_type == "in":
        stock.quantity_on_hand += quantity
    elif movement_type == "out":
        stock.quantity_on_hand -= quantity
    else:
        stock.quantity_on_hand = quantity
    stock.quantity_available = stock.quantity_on_hand - stock.quantity_reserved
    count = db.query(func.count(StockMovement.id)).scalar()
    movement = StockMovement(
        movement_number=f"SM-{(count or 0) + 1:05d}",
        product_id=product_id, warehouse_id=warehouse_id,
        movement_type=movement_type, quantity=quantity,
        notes=data.get("notes"), created_by=current_user.id
    )
    db.add(movement); db.commit()
    return {"message": "Stock adjusted", "new_quantity": stock.quantity_on_hand}

@router.get("/movements")
async def list_movements(skip: int = 0, limit: int = 50, product_id: Optional[int] = None,
                          db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    q = db.query(StockMovement)
    if product_id: q = q.filter(StockMovement.product_id == product_id)
    total = q.count()
    items = q.order_by(StockMovement.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [{"id": m.id, "movement_number": m.movement_number,
            "product_id": m.product_id, "warehouse_id": m.warehouse_id, "movement_type": m.movement_type,
            "quantity": m.quantity, "notes": m.notes, "created_at": m.created_at.isoformat() if m.created_at else None} for m in items]}
