from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from pydantic import BaseModel
import io, openpyxl
from openpyxl.styles import Font, PatternFill

from app.db.base import get_db
from app.models.asset import Asset, AssetMaintenance, AssetDepreciation
from app.models.user import User
from app.api.v1.endpoints.auth import get_current_user

router = APIRouter()

class AssetCreate(BaseModel):
    name: str
    asset_type: Optional[str] = None
    category: Optional[str] = None
    make: Optional[str] = None
    model: Optional[str] = None
    serial_number: Optional[str] = None
    purchase_date: Optional[str] = None
    purchase_price: Optional[float] = None
    warranty_expiry: Optional[str] = None
    location: Optional[str] = None
    assigned_to: Optional[int] = None
    department_id: Optional[int] = None
    depreciation_method: str = "straight_line"
    useful_life_years: Optional[int] = None
    salvage_value: float = 0

def asset_dict(a: Asset):
    return {"id": a.id, "asset_number": a.asset_number, "name": a.name, "asset_type": a.asset_type,
            "category": a.category, "make": a.make, "model": a.model, "serial_number": a.serial_number,
            "purchase_price": float(a.purchase_price or 0), "book_value": float(a.book_value or 0),
            "condition": a.condition, "status": a.status, "location": a.location,
            "assigned_to": a.assigned_to, "department_id": a.department_id, "is_active": a.is_active,
            "created_at": a.created_at.isoformat() if a.created_at else None}

@router.get("/")
async def list_assets(skip: int = 0, limit: int = 50, status: Optional[str] = None,
                       category: Optional[str] = None, db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user)):
    q = db.query(Asset).filter(Asset.is_active == True)
    if status: q = q.filter(Asset.status == status)
    if category: q = q.filter(Asset.category == category)
    total = q.count()
    items = q.order_by(Asset.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [asset_dict(a) for a in items]}

@router.post("/")
async def create_asset(data: AssetCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from datetime import datetime
    count = db.query(func.count(Asset.id)).scalar()
    asset_data = data.dict()
    if asset_data.get("purchase_date"):
        asset_data["purchase_date"] = datetime.strptime(asset_data["purchase_date"], "%Y-%m-%d").date()
    if asset_data.get("warranty_expiry"):
        asset_data["warranty_expiry"] = datetime.strptime(asset_data["warranty_expiry"], "%Y-%m-%d").date()
    a = Asset(asset_number=f"AST-{(count or 0) + 1:05d}", **asset_data)
    a.book_value = a.purchase_price
    db.add(a); db.commit(); db.refresh(a)
    return asset_dict(a)

@router.get("/export")
async def export_assets(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    items = db.query(Asset).filter(Asset.is_active == True).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Assets"
    headers = ["Asset #", "Name", "Type", "Category", "Make", "Model", "Serial #", "Purchase Price", "Book Value", "Status", "Location"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6366F1")
    for row, a in enumerate(items, 2):
        for col, val in enumerate([a.asset_number, a.name, a.asset_type, a.category, a.make, a.model,
                                    a.serial_number, float(a.purchase_price or 0), float(a.book_value or 0),
                                    a.status, a.location], 1):
            ws.cell(row=row, column=col, value=val)
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=assets.xlsx"})

@router.get("/{asset_id}")
async def get_asset(asset_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    a = db.query(Asset).filter(Asset.id == asset_id).first()
    if not a: raise HTTPException(404, "Not found")
    d = asset_dict(a)
    d["maintenances"] = [{"id": m.id, "maintenance_type": m.maintenance_type,
                          "scheduled_date": str(m.scheduled_date), "cost": float(m.cost or 0),
                          "status": m.status} for m in a.maintenances]
    return d

@router.put("/{asset_id}")
async def update_asset(asset_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    a = db.query(Asset).filter(Asset.id == asset_id).first()
    if not a: raise HTTPException(404, "Not found")
    for k, v in data.items():
        if hasattr(a, k): setattr(a, k, v)
    db.commit(); return asset_dict(a)

@router.delete("/{asset_id}")
async def delete_asset(asset_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    a = db.query(Asset).filter(Asset.id == asset_id).first()
    if not a: raise HTTPException(404, "Not found")
    a.is_active = False; db.commit(); return {"message": "Deleted"}

@router.post("/{asset_id}/maintenance")
async def schedule_maintenance(asset_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from datetime import datetime
    m = AssetMaintenance(asset_id=asset_id, maintenance_type=data.get("maintenance_type"),
                          cost=data.get("cost"), description=data.get("description"),
                          technician=data.get("technician"), status="scheduled")
    if data.get("scheduled_date"):
        m.scheduled_date = datetime.strptime(data["scheduled_date"], "%Y-%m-%d").date()
    db.add(m); db.commit(); db.refresh(m)
    return {"id": m.id, "status": m.status}
