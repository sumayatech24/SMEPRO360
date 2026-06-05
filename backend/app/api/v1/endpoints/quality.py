from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
import io, openpyxl
from openpyxl.styles import Font, PatternFill

from app.db.base import get_db
from app.models.quality import QualityCheck, QualityParameter, NonConformance
from app.models.user import User
from app.api.v1.endpoints.auth import get_current_user

router = APIRouter()

@router.get("/checks")
async def list_checks(skip: int = 0, limit: int = 50, status: Optional[str] = None,
                       check_type: Optional[str] = None, db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user)):
    q = db.query(QualityCheck).filter(QualityCheck.is_active == True)
    if status: q = q.filter(QualityCheck.status == status)
    if check_type: q = q.filter(QualityCheck.check_type == check_type)
    total = q.count()
    items = q.order_by(QualityCheck.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [{"id": c.id, "check_number": c.check_number,
            "check_type": c.check_type, "product_id": c.product_id, "batch_number": c.batch_number,
            "quantity_checked": c.quantity_checked, "quantity_passed": c.quantity_passed,
            "quantity_failed": c.quantity_failed, "status": c.status,
            "created_at": c.created_at.isoformat() if c.created_at else None} for c in items]}

@router.post("/checks")
async def create_check(data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from datetime import datetime
    count = db.query(func.count(QualityCheck.id)).scalar()
    c = QualityCheck(
        check_number=f"QC-{(count or 0) + 1:05d}",
        check_type=data.get("check_type"), product_id=data.get("product_id"),
        batch_number=data.get("batch_number"), quantity_checked=data.get("quantity_checked", 0),
        quantity_passed=data.get("quantity_passed", 0), quantity_failed=data.get("quantity_failed", 0),
        reference_type=data.get("reference_type"), reference_id=data.get("reference_id"),
        inspector_id=current_user.id, check_date=datetime.utcnow(),
        notes=data.get("notes"), results=data.get("results", [])
    )
    total_checked = float(c.quantity_checked or 0)
    total_failed = float(c.quantity_failed or 0)
    if total_checked > 0:
        if total_failed == 0: c.status = "passed"
        elif total_failed == total_checked: c.status = "failed"
        else: c.status = "conditional"
    db.add(c); db.commit(); db.refresh(c)
    return {"id": c.id, "check_number": c.check_number, "status": c.status}

@router.get("/checks/export")
async def export_checks(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    items = db.query(QualityCheck).filter(QualityCheck.is_active == True).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Quality Checks"
    headers = ["Check #", "Type", "Product ID", "Batch", "Checked", "Passed", "Failed", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6366F1")
    for row, c in enumerate(items, 2):
        for col, val in enumerate([c.check_number, c.check_type, c.product_id, c.batch_number,
                                    c.quantity_checked, c.quantity_passed, c.quantity_failed, c.status], 1):
            ws.cell(row=row, column=col, value=val)
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=quality_checks.xlsx"})

@router.get("/nonconformances")
async def list_nc(skip: int = 0, limit: int = 50, status: Optional[str] = None,
                   db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    q = db.query(NonConformance).filter(NonConformance.is_active == True)
    if status: q = q.filter(NonConformance.status == status)
    total = q.count()
    items = q.order_by(NonConformance.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [{"id": n.id, "nc_number": n.nc_number, "defect_type": n.defect_type,
            "severity": n.severity, "description": n.description, "status": n.status,
            "product_id": n.product_id} for n in items]}

@router.post("/nonconformances")
async def create_nc(data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    count = db.query(func.count(NonConformance.id)).scalar()
    nc = NonConformance(nc_number=f"NC-{(count or 0) + 1:05d}", defect_type=data.get("defect_type"),
                         severity=data.get("severity"), description=data.get("description"),
                         product_id=data.get("product_id"), raised_by=current_user.id,
                         root_cause=data.get("root_cause"), corrective_action=data.get("corrective_action"))
    db.add(nc); db.commit(); db.refresh(nc)
    return {"id": nc.id, "nc_number": nc.nc_number}

@router.get("/parameters")
async def list_parameters(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    params = db.query(QualityParameter).filter(QualityParameter.is_active == True).all()
    return [{"id": p.id, "name": p.name, "parameter_type": p.parameter_type, "unit": p.unit,
             "min_value": p.min_value, "max_value": p.max_value, "target_value": p.target_value,
             "is_critical": p.is_critical} for p in params]

@router.post("/parameters")
async def create_parameter(data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    p = QualityParameter(**{k: v for k, v in data.items() if hasattr(QualityParameter, k)})
    db.add(p); db.commit(); db.refresh(p)
    return {"id": p.id, "name": p.name}
