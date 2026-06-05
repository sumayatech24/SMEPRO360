from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from pydantic import BaseModel
import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.db.base import get_db
from app.models.crm import Customer, Contact, Opportunity, Activity
from app.models.user import User
from app.api.v1.endpoints.auth import get_current_user

router = APIRouter()

# ─── Customers ───────────────────────────────────────────────────────────────

class CustomerCreate(BaseModel):
    company_name: str
    customer_type: str = "b2b"
    industry: Optional[str] = None
    gstin: Optional[str] = None
    credit_limit: float = 0
    credit_days: int = 30
    billing_address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    segment: Optional[str] = None
    account_manager_id: Optional[int] = None

def cust_dict(c: Customer):
    return {"id": c.id, "customer_number": c.customer_number, "company_name": c.company_name,
            "customer_type": c.customer_type, "industry": c.industry, "gstin": c.gstin,
            "credit_limit": float(c.credit_limit or 0), "credit_days": c.credit_days,
            "city": c.city, "state": c.state, "phone": c.phone, "email": c.email,
            "status": c.status, "segment": c.segment, "is_active": c.is_active,
            "created_at": c.created_at.isoformat() if c.created_at else None}

@router.get("/customers")
async def list_customers(skip: int = 0, limit: int = 50, search: Optional[str] = None,
                          status: Optional[str] = None, db: Session = Depends(get_db),
                          current_user: User = Depends(get_current_user)):
    q = db.query(Customer).filter(Customer.is_active == True)
    if search:
        q = q.filter(Customer.company_name.ilike(f"%{search}%") | Customer.email.ilike(f"%{search}%"))
    if status:
        q = q.filter(Customer.status == status)
    total = q.count()
    items = q.order_by(Customer.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [cust_dict(c) for c in items]}

@router.post("/customers")
async def create_customer(data: CustomerCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    count = db.query(func.count(Customer.id)).scalar()
    c = Customer(customer_number=f"CUS-{(count or 0) + 1:05d}", **data.dict())
    db.add(c); db.commit(); db.refresh(c)
    return cust_dict(c)

@router.get("/customers/stats")
async def customer_stats(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    total = db.query(func.count(Customer.id)).filter(Customer.is_active == True).scalar()
    by_type = db.query(Customer.customer_type, func.count(Customer.id)).group_by(Customer.customer_type).all()
    by_industry = db.query(Customer.industry, func.count(Customer.id)).filter(Customer.is_active == True).group_by(Customer.industry).limit(10).all()
    return {"total": total, "by_type": dict(by_type), "by_industry": {k: v for k, v in by_industry if k}}

@router.get("/customers/export")
async def export_customers(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    items = db.query(Customer).filter(Customer.is_active == True).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Customers"
    headers = ["Customer #", "Company", "Type", "Industry", "GSTIN", "City", "State", "Phone", "Email", "Status", "Credit Limit"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6366F1")
    for row, c in enumerate(items, 2):
        for col, val in enumerate([c.customer_number, c.company_name, c.customer_type, c.industry,
                                    c.gstin, c.city, c.state, c.phone, c.email, c.status, float(c.credit_limit or 0)], 1):
            ws.cell(row=row, column=col, value=val)
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=customers.xlsx"})

@router.get("/customers/{customer_id}")
async def get_customer(customer_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    c = db.query(Customer).filter(Customer.id == customer_id).first()
    if not c: raise HTTPException(404, "Customer not found")
    return cust_dict(c)

@router.put("/customers/{customer_id}")
async def update_customer(customer_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    c = db.query(Customer).filter(Customer.id == customer_id).first()
    if not c: raise HTTPException(404, "Customer not found")
    for k, v in data.items():
        if hasattr(c, k): setattr(c, k, v)
    db.commit(); return cust_dict(c)

@router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    c = db.query(Customer).filter(Customer.id == customer_id).first()
    if not c: raise HTTPException(404, "Not found")
    c.is_active = False; db.commit(); return {"message": "Deleted"}

# ─── Contacts ───────────────────────────────────────────────────────────────

class ContactCreate(BaseModel):
    customer_id: int
    first_name: str
    last_name: Optional[str] = None
    designation: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    is_primary: bool = False

@router.get("/contacts")
async def list_contacts(customer_id: Optional[int] = None, skip: int = 0, limit: int = 50,
                         db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    q = db.query(Contact).filter(Contact.is_active == True)
    if customer_id: q = q.filter(Contact.customer_id == customer_id)
    total = q.count()
    items = q.offset(skip).limit(limit).all()
    return {"total": total, "items": [{"id": c.id, "customer_id": c.customer_id, "first_name": c.first_name,
            "last_name": c.last_name, "designation": c.designation, "email": c.email, "phone": c.phone,
            "is_primary": c.is_primary} for c in items]}

@router.post("/contacts")
async def create_contact(data: ContactCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    c = Contact(**data.dict()); db.add(c); db.commit(); db.refresh(c)
    return {"id": c.id, "first_name": c.first_name}

@router.put("/contacts/{contact_id}")
async def update_contact(contact_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    c = db.query(Contact).filter(Contact.id == contact_id).first()
    if not c: raise HTTPException(404, "Not found")
    for k, v in data.items():
        if hasattr(c, k): setattr(c, k, v)
    db.commit(); return {"id": c.id}

@router.delete("/contacts/{contact_id}")
async def delete_contact(contact_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    c = db.query(Contact).filter(Contact.id == contact_id).first()
    if not c: raise HTTPException(404, "Not found")
    c.is_active = False; db.commit(); return {"message": "Deleted"}

# ─── Opportunities ──────────────────────────────────────────────────────────

class OpportunityCreate(BaseModel):
    name: str
    customer_id: int
    stage: str = "prospecting"
    probability: float = 10
    expected_revenue: Optional[float] = None
    assigned_to: Optional[int] = None
    description: Optional[str] = None

def opp_dict(o: Opportunity):
    return {"id": o.id, "opportunity_number": o.opportunity_number, "name": o.name,
            "customer_id": o.customer_id, "stage": o.stage, "probability": o.probability,
            "expected_revenue": float(o.expected_revenue or 0), "assigned_to": o.assigned_to,
            "description": o.description, "is_active": o.is_active,
            "created_at": o.created_at.isoformat() if o.created_at else None}

@router.get("/opportunities")
async def list_opportunities(skip: int = 0, limit: int = 50, stage: Optional[str] = None,
                              customer_id: Optional[int] = None,
                              db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    q = db.query(Opportunity).filter(Opportunity.is_active == True)
    if stage: q = q.filter(Opportunity.stage == stage)
    if customer_id: q = q.filter(Opportunity.customer_id == customer_id)
    total = q.count()
    items = q.order_by(Opportunity.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [opp_dict(o) for o in items]}

@router.post("/opportunities")
async def create_opportunity(data: OpportunityCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    count = db.query(func.count(Opportunity.id)).scalar()
    o = Opportunity(opportunity_number=f"OPP-{(count or 0) + 1:05d}", **data.dict())
    db.add(o); db.commit(); db.refresh(o)
    return opp_dict(o)

@router.get("/opportunities/pipeline")
async def pipeline_view(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    stages = ["prospecting", "qualification", "proposal", "negotiation", "closed_won", "closed_lost"]
    result = {}
    for stage in stages:
        q = db.query(func.count(Opportunity.id), func.sum(Opportunity.expected_revenue)).filter(
            Opportunity.stage == stage, Opportunity.is_active == True).first()
        result[stage] = {"count": q[0] or 0, "value": float(q[1] or 0)}
    return result

@router.get("/opportunities/export")
async def export_opportunities(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    items = db.query(Opportunity).filter(Opportunity.is_active == True).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Opportunities"
    headers = ["Opp #", "Name", "Customer ID", "Stage", "Probability %", "Expected Revenue"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6366F1")
    for row, o in enumerate(items, 2):
        for col, val in enumerate([o.opportunity_number, o.name, o.customer_id, o.stage,
                                    o.probability, float(o.expected_revenue or 0)], 1):
            ws.cell(row=row, column=col, value=val)
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=opportunities.xlsx"})

@router.put("/opportunities/{opp_id}")
async def update_opportunity(opp_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    o = db.query(Opportunity).filter(Opportunity.id == opp_id).first()
    if not o: raise HTTPException(404, "Not found")
    for k, v in data.items():
        if hasattr(o, k): setattr(o, k, v)
    db.commit(); return opp_dict(o)

@router.delete("/opportunities/{opp_id}")
async def delete_opportunity(opp_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    o = db.query(Opportunity).filter(Opportunity.id == opp_id).first()
    if not o: raise HTTPException(404, "Not found")
    o.is_active = False; db.commit(); return {"message": "Deleted"}
