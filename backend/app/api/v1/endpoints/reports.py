from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, timedelta
import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.db.base import get_db
from app.models.user import User
from app.models.sales import SalesOrder, Invoice
from app.models.lead import Lead
from app.models.crm import Customer
from app.models.hr import Employee, Payroll
from app.models.inventory import StockLevel, Product
from app.api.v1.endpoints.auth import get_current_user

router = APIRouter()

@router.get("/sales-summary")
async def sales_summary(days: int = 30, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    since = datetime.utcnow() - timedelta(days=days)
    total_orders = db.query(func.count(SalesOrder.id)).filter(SalesOrder.created_at >= since, SalesOrder.is_active == True).scalar() or 0
    total_revenue = db.query(func.sum(SalesOrder.total_amount)).filter(SalesOrder.created_at >= since, SalesOrder.is_active == True).scalar() or 0
    by_status = dict(db.query(SalesOrder.status, func.count(SalesOrder.id)).filter(SalesOrder.created_at >= since).group_by(SalesOrder.status).all())
    return {"period_days": days, "total_orders": total_orders, "total_revenue": float(total_revenue), "by_status": by_status}

@router.get("/lead-funnel")
async def lead_funnel(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    statuses = ["new", "contacted", "qualified", "converted"]
    result = {}
    for s in statuses:
        count = db.query(func.count(Lead.id)).filter(Lead.status == s, Lead.is_active == True).scalar() or 0
        result[s] = count
    return result

@router.get("/inventory-valuation")
async def inventory_valuation(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    stocks = db.query(StockLevel).all()
    total_value = 0
    items = []
    for s in stocks:
        product = db.query(Product).filter(Product.id == s.product_id).first()
        if product:
            value = s.quantity_on_hand * float(product.cost_price or 0)
            total_value += value
            items.append({"product_id": s.product_id, "sku": product.sku, "name": product.name,
                          "quantity": s.quantity_on_hand, "cost_price": float(product.cost_price or 0),
                          "total_value": value})
    return {"total_value": total_value, "items": items}

@router.get("/hr-summary")
async def hr_summary(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    total_employees = db.query(func.count(Employee.id)).filter(Employee.is_active == True).scalar() or 0
    by_dept = dict(db.query(Employee.department_id, func.count(Employee.id)).filter(Employee.is_active == True).group_by(Employee.department_id).all())
    by_type = dict(db.query(Employee.employment_type, func.count(Employee.id)).filter(Employee.is_active == True).group_by(Employee.employment_type).all())
    total_payroll = db.query(func.sum(Payroll.net_salary)).scalar() or 0
    return {"total_employees": total_employees, "by_department": by_dept, "by_type": by_type,
            "total_payroll_disbursed": float(total_payroll)}

@router.get("/export/master-report")
async def export_master_report(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    wb = openpyxl.Workbook()

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill("solid", start_color="6366F1")
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = 18

    # Sales sheet
    ws1 = wb.active; ws1.title = "Sales Orders"
    headers1 = ["Order #", "Customer ID", "Status", "Date", "Total Amount", "Payment Status"]
    style_header(ws1, headers1)
    orders = db.query(SalesOrder).filter(SalesOrder.is_active == True).all()
    for row, o in enumerate(orders, 2):
        for col, val in enumerate([o.order_number, o.customer_id, o.status,
                                    o.order_date.strftime("%Y-%m-%d") if o.order_date else "",
                                    float(o.total_amount or 0), o.payment_status], 1):
            ws1.cell(row=row, column=col, value=val)

    # Customers sheet
    from app.models.crm import Customer
    ws2 = wb.create_sheet("Customers")
    headers2 = ["Customer #", "Company", "Industry", "City", "Phone", "Email", "Status"]
    style_header(ws2, headers2)
    customers = db.query(Customer).filter(Customer.is_active == True).all()
    for row, c in enumerate(customers, 2):
        for col, val in enumerate([c.customer_number, c.company_name, c.industry, c.city, c.phone, c.email, c.status], 1):
            ws2.cell(row=row, column=col, value=val)

    # Employees sheet
    ws3 = wb.create_sheet("Employees")
    headers3 = ["Emp #", "First Name", "Last Name", "Email", "Department ID", "Basic Salary", "Status"]
    style_header(ws3, headers3)
    employees = db.query(Employee).filter(Employee.is_active == True).all()
    for row, e in enumerate(employees, 2):
        for col, val in enumerate([e.employee_number, e.first_name, e.last_name, e.email,
                                    e.department_id, float(e.basic_salary or 0), e.status], 1):
            ws3.cell(row=row, column=col, value=val)

    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=SMEPRO360_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"})
