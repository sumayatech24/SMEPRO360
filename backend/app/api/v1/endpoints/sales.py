from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional, List
from pydantic import BaseModel
from decimal import Decimal
import io, openpyxl
from openpyxl.styles import Font, PatternFill

from app.db.base import get_db
from app.models.sales import SalesOrder, SalesOrderItem, Invoice, InvoiceItem, Payment
from app.models.user import User
from app.api.v1.endpoints.auth import get_current_user

router = APIRouter()

class OrderItemIn(BaseModel):
    product_id: Optional[int] = None
    description: str
    quantity: float
    unit: Optional[str] = "nos"
    unit_price: float
    discount_percent: float = 0
    tax_percent: float = 18
    hsn_code: Optional[str] = None

class SalesOrderCreate(BaseModel):
    customer_id: int
    delivery_date: Optional[str] = None
    payment_terms: Optional[str] = None
    notes: Optional[str] = None
    assigned_to: Optional[int] = None
    purchase_order_number: Optional[str] = None
    items: List[OrderItemIn] = []

def so_dict(o: SalesOrder):
    return {"id": o.id, "order_number": o.order_number, "customer_id": o.customer_id,
            "status": o.status, "order_date": o.order_date.isoformat() if o.order_date else None,
            "subtotal": float(o.subtotal or 0), "tax_amount": float(o.tax_amount or 0),
            "total_amount": float(o.total_amount or 0), "payment_status": o.payment_status,
            "notes": o.notes, "assigned_to": o.assigned_to,
            "created_at": o.created_at.isoformat() if o.created_at else None}

@router.get("/orders")
async def list_orders(skip: int = 0, limit: int = 50, status: Optional[str] = None,
                       customer_id: Optional[int] = None, db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user)):
    q = db.query(SalesOrder).filter(SalesOrder.is_active == True)
    if status: q = q.filter(SalesOrder.status == status)
    if customer_id: q = q.filter(SalesOrder.customer_id == customer_id)
    total = q.count()
    items = q.order_by(SalesOrder.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [so_dict(o) for o in items]}

@router.post("/orders")
async def create_order(data: SalesOrderCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    count = db.query(func.count(SalesOrder.id)).scalar()
    order = SalesOrder(
        order_number=f"SO-{(count or 0) + 1:05d}",
        customer_id=data.customer_id, payment_terms=data.payment_terms,
        notes=data.notes, assigned_to=data.assigned_to,
        purchase_order_number=data.purchase_order_number
    )
    subtotal = tax_total = 0
    for item_data in data.items:
        line = item_data.quantity * item_data.unit_price * (1 - item_data.discount_percent / 100)
        tax = line * item_data.tax_percent / 100
        item = SalesOrderItem(
            product_id=item_data.product_id, description=item_data.description,
            quantity=item_data.quantity, unit=item_data.unit, unit_price=item_data.unit_price,
            discount_percent=item_data.discount_percent, tax_percent=item_data.tax_percent,
            tax_amount=tax, line_total=line + tax, hsn_code=item_data.hsn_code
        )
        order.items.append(item)
        subtotal += line; tax_total += tax
    order.subtotal = subtotal; order.tax_amount = tax_total; order.total_amount = subtotal + tax_total
    db.add(order); db.commit(); db.refresh(order)
    return so_dict(order)

@router.get("/orders/stats")
async def order_stats(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    total = db.query(func.count(SalesOrder.id)).filter(SalesOrder.is_active == True).scalar()
    revenue = db.query(func.sum(SalesOrder.total_amount)).filter(SalesOrder.status == "confirmed").scalar()
    by_status = db.query(SalesOrder.status, func.count(SalesOrder.id)).filter(SalesOrder.is_active == True).group_by(SalesOrder.status).all()
    return {"total": total, "total_revenue": float(revenue or 0), "by_status": dict(by_status)}

@router.get("/orders/export")
async def export_orders(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    items = db.query(SalesOrder).filter(SalesOrder.is_active == True).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Sales Orders"
    headers = ["Order #", "Customer ID", "Status", "Order Date", "Subtotal", "Tax", "Total", "Payment Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6366F1")
    for row, o in enumerate(items, 2):
        for col, val in enumerate([o.order_number, o.customer_id, o.status,
                                    o.order_date.strftime("%Y-%m-%d") if o.order_date else "",
                                    float(o.subtotal or 0), float(o.tax_amount or 0),
                                    float(o.total_amount or 0), o.payment_status], 1):
            ws.cell(row=row, column=col, value=val)
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=sales_orders.xlsx"})

@router.get("/orders/{order_id}")
async def get_order(order_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    o = db.query(SalesOrder).filter(SalesOrder.id == order_id).first()
    if not o: raise HTTPException(404, "Not found")
    d = so_dict(o)
    d["items"] = [{"id": i.id, "description": i.description, "quantity": i.quantity,
                   "unit_price": float(i.unit_price), "tax_percent": i.tax_percent,
                   "line_total": float(i.line_total or 0)} for i in o.items]
    return d

@router.put("/orders/{order_id}")
async def update_order(order_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    o = db.query(SalesOrder).filter(SalesOrder.id == order_id).first()
    if not o: raise HTTPException(404, "Not found")
    for k, v in data.items():
        if hasattr(o, k): setattr(o, k, v)
    db.commit(); return so_dict(o)

@router.delete("/orders/{order_id}")
async def delete_order(order_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    o = db.query(SalesOrder).filter(SalesOrder.id == order_id).first()
    if not o: raise HTTPException(404, "Not found")
    o.is_active = False; db.commit(); return {"message": "Deleted"}

# ─── Invoices ────────────────────────────────────────────────────────────────

class InvoiceCreate(BaseModel):
    customer_id: int
    sales_order_id: Optional[int] = None
    due_date: Optional[str] = None
    notes: Optional[str] = None
    items: List[OrderItemIn] = []

def inv_dict(inv: Invoice):
    return {"id": inv.id, "invoice_number": inv.invoice_number, "customer_id": inv.customer_id,
            "status": inv.status, "invoice_date": inv.invoice_date.isoformat() if inv.invoice_date else None,
            "due_date": inv.due_date.isoformat() if inv.due_date else None,
            "total_amount": float(inv.total_amount or 0), "amount_paid": float(inv.amount_paid or 0),
            "balance_due": float(inv.balance_due or 0)}

@router.get("/invoices")
async def list_invoices(skip: int = 0, limit: int = 50, status: Optional[str] = None,
                         customer_id: Optional[int] = None, db: Session = Depends(get_db),
                         current_user: User = Depends(get_current_user)):
    q = db.query(Invoice).filter(Invoice.is_active == True)
    if status: q = q.filter(Invoice.status == status)
    if customer_id: q = q.filter(Invoice.customer_id == customer_id)
    total = q.count()
    items = q.order_by(Invoice.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [inv_dict(i) for i in items]}

@router.post("/invoices")
async def create_invoice(data: InvoiceCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    count = db.query(func.count(Invoice.id)).scalar()
    inv = Invoice(
        invoice_number=f"INV-{(count or 0) + 1:05d}",
        customer_id=data.customer_id, sales_order_id=data.sales_order_id, notes=data.notes
    )
    subtotal = tax_total = 0
    for item_data in data.items:
        line = item_data.quantity * item_data.unit_price * (1 - item_data.discount_percent / 100)
        cgst = line * (item_data.tax_percent / 2) / 100
        sgst = line * (item_data.tax_percent / 2) / 100
        item = InvoiceItem(
            product_id=item_data.product_id, description=item_data.description,
            quantity=item_data.quantity, unit_price=item_data.unit_price,
            cgst_percent=item_data.tax_percent / 2, sgst_percent=item_data.tax_percent / 2,
            line_total=line + cgst + sgst, hsn_code=item_data.hsn_code
        )
        inv.items.append(item)
        subtotal += line; tax_total += cgst + sgst
    inv.subtotal = subtotal; inv.cgst_amount = tax_total / 2; inv.sgst_amount = tax_total / 2
    inv.total_amount = subtotal + tax_total; inv.balance_due = inv.total_amount
    db.add(inv); db.commit(); db.refresh(inv)
    return inv_dict(inv)

@router.get("/invoices/export")
async def export_invoices(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    items = db.query(Invoice).filter(Invoice.is_active == True).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Invoices"
    headers = ["Invoice #", "Customer ID", "Status", "Date", "Total", "Paid", "Balance Due"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6366F1")
    for row, inv in enumerate(items, 2):
        for col, val in enumerate([inv.invoice_number, inv.customer_id, inv.status,
                                    inv.invoice_date.strftime("%Y-%m-%d") if inv.invoice_date else "",
                                    float(inv.total_amount or 0), float(inv.amount_paid or 0),
                                    float(inv.balance_due or 0)], 1):
            ws.cell(row=row, column=col, value=val)
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=invoices.xlsx"})

@router.get("/invoices/{inv_id}")
async def get_invoice(inv_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    inv = db.query(Invoice).filter(Invoice.id == inv_id).first()
    if not inv: raise HTTPException(404, "Not found")
    d = inv_dict(inv)
    d["items"] = [{"id": i.id, "description": i.description, "quantity": i.quantity,
                   "unit_price": float(i.unit_price), "line_total": float(i.line_total or 0)} for i in inv.items]
    return d

@router.put("/invoices/{inv_id}")
async def update_invoice(inv_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    inv = db.query(Invoice).filter(Invoice.id == inv_id).first()
    if not inv: raise HTTPException(404, "Not found")
    for k, v in data.items():
        if hasattr(inv, k): setattr(inv, k, v)
    db.commit(); return inv_dict(inv)

# ─── Payments ────────────────────────────────────────────────────────────────

class PaymentCreate(BaseModel):
    invoice_id: int
    customer_id: int
    amount: float
    payment_mode: str = "bank_transfer"
    reference_number: Optional[str] = None
    notes: Optional[str] = None

@router.get("/payments")
async def list_payments(skip: int = 0, limit: int = 50, db: Session = Depends(get_db),
                         current_user: User = Depends(get_current_user)):
    q = db.query(Payment)
    total = q.count()
    items = q.order_by(Payment.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [{"id": p.id, "payment_number": p.payment_number,
            "invoice_id": p.invoice_id, "amount": float(p.amount), "payment_mode": p.payment_mode,
            "reference_number": p.reference_number, "status": p.status} for p in items]}

@router.post("/payments")
async def create_payment(data: PaymentCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from datetime import datetime
    count = db.query(func.count(Payment.id)).scalar()
    payment = Payment(
        payment_number=f"PAY-{(count or 0) + 1:05d}",
        invoice_id=data.invoice_id, customer_id=data.customer_id,
        amount=data.amount, payment_mode=data.payment_mode,
        reference_number=data.reference_number, notes=data.notes,
        payment_date=datetime.utcnow()
    )
    db.add(payment)
    inv = db.query(Invoice).filter(Invoice.id == data.invoice_id).first()
    if inv:
        inv.amount_paid = float(inv.amount_paid or 0) + data.amount
        inv.balance_due = float(inv.total_amount or 0) - float(inv.amount_paid or 0)
        if float(inv.balance_due) <= 0:
            inv.status = "paid"
    db.commit(); db.refresh(payment)
    return {"id": payment.id, "payment_number": payment.payment_number}
