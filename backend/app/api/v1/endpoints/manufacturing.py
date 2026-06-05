from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional, List
from pydantic import BaseModel
import io, openpyxl
from openpyxl.styles import Font, PatternFill

from app.db.base import get_db
from app.models.manufacturing import BOM, BOMItem, WorkOrder, WorkOrderOperation, ProductionEntry
from app.models.user import User
from app.api.v1.endpoints.auth import get_current_user

router = APIRouter()

class BOMItemIn(BaseModel):
    product_id: int
    quantity: float
    unit: Optional[str] = "nos"
    scrap_percent: float = 0

class BOMCreate(BaseModel):
    name: str
    product_id: int
    quantity: float = 1
    unit: Optional[str] = "nos"
    version: str = "1.0"
    notes: Optional[str] = None
    items: List[BOMItemIn] = []

def bom_dict(b: BOM):
    return {"id": b.id, "bom_number": b.bom_number, "name": b.name, "product_id": b.product_id,
            "quantity": b.quantity, "version": b.version, "status": b.status, "is_active": b.is_active,
            "created_at": b.created_at.isoformat() if b.created_at else None}

@router.get("/boms")
async def list_boms(skip: int = 0, limit: int = 50, product_id: Optional[int] = None,
                     db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    q = db.query(BOM).filter(BOM.is_active == True)
    if product_id: q = q.filter(BOM.product_id == product_id)
    total = q.count()
    items = q.order_by(BOM.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [bom_dict(b) for b in items]}

@router.post("/boms")
async def create_bom(data: BOMCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    count = db.query(func.count(BOM.id)).scalar()
    bom = BOM(bom_number=f"BOM-{(count or 0) + 1:05d}", name=data.name, product_id=data.product_id,
               quantity=data.quantity, unit=data.unit, version=data.version, notes=data.notes)
    for item_data in data.items:
        item = BOMItem(**item_data.dict()); bom.items.append(item)
    db.add(bom); db.commit(); db.refresh(bom)
    return bom_dict(bom)

@router.get("/boms/{bom_id}")
async def get_bom(bom_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    b = db.query(BOM).filter(BOM.id == bom_id).first()
    if not b: raise HTTPException(404, "Not found")
    d = bom_dict(b)
    d["items"] = [{"id": i.id, "product_id": i.product_id, "quantity": i.quantity,
                   "unit": i.unit, "scrap_percent": i.scrap_percent} for i in b.items]
    return d

@router.put("/boms/{bom_id}")
async def update_bom(bom_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    b = db.query(BOM).filter(BOM.id == bom_id).first()
    if not b: raise HTTPException(404, "Not found")
    for k, v in data.items():
        if hasattr(b, k): setattr(b, k, v)
    db.commit(); return bom_dict(b)

# ─── Work Orders ──────────────────────────────────────────────────────────────

class WOCreate(BaseModel):
    product_id: int
    bom_id: Optional[int] = None
    planned_quantity: float
    priority: str = "normal"
    planned_start: Optional[str] = None
    planned_end: Optional[str] = None
    warehouse_id: Optional[int] = None
    notes: Optional[str] = None

def wo_dict(w: WorkOrder):
    return {"id": w.id, "wo_number": w.wo_number, "product_id": w.product_id, "bom_id": w.bom_id,
            "planned_quantity": w.planned_quantity, "produced_quantity": w.produced_quantity,
            "rejected_quantity": w.rejected_quantity, "status": w.status, "priority": w.priority,
            "planned_start": w.planned_start.isoformat() if w.planned_start else None,
            "planned_end": w.planned_end.isoformat() if w.planned_end else None,
            "is_active": w.is_active, "created_at": w.created_at.isoformat() if w.created_at else None}

@router.get("/workorders")
async def list_workorders(skip: int = 0, limit: int = 50, status: Optional[str] = None,
                           db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    q = db.query(WorkOrder).filter(WorkOrder.is_active == True)
    if status: q = q.filter(WorkOrder.status == status)
    total = q.count()
    items = q.order_by(WorkOrder.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [wo_dict(w) for w in items]}

@router.post("/workorders")
async def create_workorder(data: WOCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from datetime import datetime
    count = db.query(func.count(WorkOrder.id)).scalar()
    wo = WorkOrder(wo_number=f"WO-{(count or 0) + 1:05d}", product_id=data.product_id,
                    bom_id=data.bom_id, planned_quantity=data.planned_quantity,
                    priority=data.priority, warehouse_id=data.warehouse_id, notes=data.notes,
                    created_by=current_user.id)
    if data.planned_start: wo.planned_start = datetime.fromisoformat(data.planned_start)
    if data.planned_end: wo.planned_end = datetime.fromisoformat(data.planned_end)
    db.add(wo); db.commit(); db.refresh(wo)
    return wo_dict(wo)

@router.get("/workorders/export")
async def export_workorders(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    items = db.query(WorkOrder).filter(WorkOrder.is_active == True).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Work Orders"
    headers = ["WO #", "Product ID", "Planned Qty", "Produced Qty", "Rejected Qty", "Status", "Priority"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6366F1")
    for row, w in enumerate(items, 2):
        for col, val in enumerate([w.wo_number, w.product_id, w.planned_quantity,
                                    w.produced_quantity, w.rejected_quantity, w.status, w.priority], 1):
            ws.cell(row=row, column=col, value=val)
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=work_orders.xlsx"})

@router.get("/workorders/{wo_id}")
async def get_workorder(wo_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    w = db.query(WorkOrder).filter(WorkOrder.id == wo_id).first()
    if not w: raise HTTPException(404, "Not found")
    d = wo_dict(w)
    d["operations"] = [{"id": o.id, "operation_name": o.operation_name, "work_center": o.work_center,
                        "sequence": o.sequence, "planned_hours": o.planned_hours, "status": o.status} for o in w.operations]
    return d

@router.put("/workorders/{wo_id}")
async def update_workorder(wo_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    w = db.query(WorkOrder).filter(WorkOrder.id == wo_id).first()
    if not w: raise HTTPException(404, "Not found")
    for k, v in data.items():
        if hasattr(w, k): setattr(w, k, v)
    db.commit(); return wo_dict(w)

@router.post("/workorders/{wo_id}/production")
async def log_production(wo_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    w = db.query(WorkOrder).filter(WorkOrder.id == wo_id).first()
    if not w: raise HTTPException(404, "Not found")
    qty_produced = data.get("quantity_produced", 0)
    qty_rejected = data.get("quantity_rejected", 0)
    entry = ProductionEntry(work_order_id=wo_id, quantity_produced=qty_produced,
                             quantity_rejected=qty_rejected, operator_id=current_user.id,
                             shift=data.get("shift"), notes=data.get("notes"),
                             rejection_reason=data.get("rejection_reason"))
    db.add(entry)
    w.produced_quantity += qty_produced; w.rejected_quantity += qty_rejected
    if w.produced_quantity >= w.planned_quantity: w.status = "completed"
    else: w.status = "in_progress"
    db.commit()
    return {"message": "Production logged", "total_produced": w.produced_quantity}
