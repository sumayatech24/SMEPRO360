from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from pydantic import BaseModel
from datetime import date
import io, openpyxl
from openpyxl.styles import Font, PatternFill

from app.db.base import get_db
from app.models.hr import Employee, Department, Designation, Attendance, Leave, Payroll
from app.models.user import User
from app.api.v1.endpoints.auth import get_current_user

router = APIRouter()

class EmployeeCreate(BaseModel):
    first_name: str
    last_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    department_id: Optional[int] = None
    designation_id: Optional[int] = None
    employment_type: str = "full_time"
    date_of_joining: Optional[str] = None
    basic_salary: float = 0
    hra: float = 0
    other_allowances: float = 0
    pan_number: Optional[str] = None
    aadhaar_number: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    bank_name: Optional[str] = None
    bank_account: Optional[str] = None
    bank_ifsc: Optional[str] = None

def emp_dict(e: Employee):
    return {"id": e.id, "employee_number": e.employee_number, "first_name": e.first_name,
            "last_name": e.last_name, "email": e.email, "phone": e.phone,
            "department_id": e.department_id, "designation_id": e.designation_id,
            "employment_type": e.employment_type,
            "date_of_joining": str(e.date_of_joining) if e.date_of_joining else None,
            "basic_salary": float(e.basic_salary or 0), "status": e.status, "is_active": e.is_active,
            "created_at": e.created_at.isoformat() if e.created_at else None}

@router.get("/employees")
async def list_employees(skip: int = 0, limit: int = 50, search: Optional[str] = None,
                          department_id: Optional[int] = None, status: Optional[str] = None,
                          db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    q = db.query(Employee).filter(Employee.is_active == True)
    if search: q = q.filter(Employee.first_name.ilike(f"%{search}%") | Employee.last_name.ilike(f"%{search}%"))
    if department_id: q = q.filter(Employee.department_id == department_id)
    if status: q = q.filter(Employee.status == status)
    total = q.count()
    items = q.order_by(Employee.first_name).offset(skip).limit(limit).all()
    return {"total": total, "items": [emp_dict(e) for e in items]}

@router.post("/employees")
async def create_employee(data: EmployeeCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    count = db.query(func.count(Employee.id)).scalar()
    emp = Employee(employee_number=f"EMP-{(count or 0) + 1:05d}", **data.dict())
    db.add(emp); db.commit(); db.refresh(emp)
    return emp_dict(emp)

@router.get("/employees/export")
async def export_employees(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    items = db.query(Employee).filter(Employee.is_active == True).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Employees"
    headers = ["Emp #", "First Name", "Last Name", "Email", "Phone", "Dept ID", "Type", "Basic Salary", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6366F1")
    for row, e in enumerate(items, 2):
        for col, val in enumerate([e.employee_number, e.first_name, e.last_name, e.email, e.phone,
                                    e.department_id, e.employment_type, float(e.basic_salary or 0), e.status], 1):
            ws.cell(row=row, column=col, value=val)
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=employees.xlsx"})

@router.get("/employees/{emp_id}")
async def get_employee(emp_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    e = db.query(Employee).filter(Employee.id == emp_id).first()
    if not e: raise HTTPException(404, "Not found")
    return emp_dict(e)

@router.put("/employees/{emp_id}")
async def update_employee(emp_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    e = db.query(Employee).filter(Employee.id == emp_id).first()
    if not e: raise HTTPException(404, "Not found")
    for k, v in data.items():
        if hasattr(e, k): setattr(e, k, v)
    db.commit(); return emp_dict(e)

@router.delete("/employees/{emp_id}")
async def delete_employee(emp_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    e = db.query(Employee).filter(Employee.id == emp_id).first()
    if not e: raise HTTPException(404, "Not found")
    e.is_active = False; db.commit(); return {"message": "Deleted"}

# ─── Departments ────────────────────────────────────────────────────────────

@router.get("/departments")
async def list_departments(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    depts = db.query(Department).filter(Department.is_active == True).all()
    return [{"id": d.id, "name": d.name, "code": d.code, "parent_id": d.parent_id} for d in depts]

@router.post("/departments")
async def create_department(data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    d = Department(**{k: v for k, v in data.items() if hasattr(Department, k)})
    db.add(d); db.commit(); db.refresh(d)
    return {"id": d.id, "name": d.name}

# ─── Attendance ──────────────────────────────────────────────────────────────

class AttendanceCreate(BaseModel):
    employee_id: int
    date: str
    check_in: Optional[str] = None
    check_out: Optional[str] = None
    status: str = "present"
    notes: Optional[str] = None

@router.get("/attendance")
async def list_attendance(employee_id: Optional[int] = None, skip: int = 0, limit: int = 100,
                           db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    q = db.query(Attendance)
    if employee_id: q = q.filter(Attendance.employee_id == employee_id)
    total = q.count()
    items = q.order_by(Attendance.date.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [{"id": a.id, "employee_id": a.employee_id, "date": str(a.date),
            "status": a.status, "hours_worked": a.hours_worked, "notes": a.notes} for a in items]}

@router.post("/attendance")
async def mark_attendance(data: AttendanceCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from datetime import datetime
    a = Attendance(employee_id=data.employee_id, status=data.status, notes=data.notes,
                   date=datetime.strptime(data.date, "%Y-%m-%d").date())
    if data.check_in: a.check_in = datetime.fromisoformat(data.check_in)
    if data.check_out: a.check_out = datetime.fromisoformat(data.check_out)
    if a.check_in and a.check_out:
        a.hours_worked = (a.check_out - a.check_in).seconds / 3600
    db.add(a); db.commit(); db.refresh(a)
    return {"id": a.id, "status": a.status}

@router.get("/attendance/export")
async def export_attendance(employee_id: Optional[int] = None, db: Session = Depends(get_db),
                             current_user: User = Depends(get_current_user)):
    q = db.query(Attendance)
    if employee_id: q = q.filter(Attendance.employee_id == employee_id)
    items = q.order_by(Attendance.date.desc()).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Attendance"
    headers = ["Employee ID", "Date", "Status", "Check In", "Check Out", "Hours Worked"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6366F1")
    for row, a in enumerate(items, 2):
        for col, val in enumerate([a.employee_id, str(a.date), a.status,
                                    str(a.check_in) if a.check_in else "",
                                    str(a.check_out) if a.check_out else "",
                                    a.hours_worked], 1):
            ws.cell(row=row, column=col, value=val)
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=attendance.xlsx"})

# ─── Leave ───────────────────────────────────────────────────────────────────

class LeaveCreate(BaseModel):
    employee_id: int
    leave_type: str
    from_date: str
    to_date: str
    reason: Optional[str] = None

@router.get("/leaves")
async def list_leaves(employee_id: Optional[int] = None, status: Optional[str] = None,
                       skip: int = 0, limit: int = 50, db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user)):
    q = db.query(Leave)
    if employee_id: q = q.filter(Leave.employee_id == employee_id)
    if status: q = q.filter(Leave.status == status)
    total = q.count()
    items = q.order_by(Leave.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [{"id": l.id, "employee_id": l.employee_id, "leave_type": l.leave_type,
            "from_date": str(l.from_date), "to_date": str(l.to_date), "days": l.days,
            "status": l.status, "reason": l.reason} for l in items]}

@router.post("/leaves")
async def apply_leave(data: LeaveCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from datetime import datetime
    fd = datetime.strptime(data.from_date, "%Y-%m-%d").date()
    td = datetime.strptime(data.to_date, "%Y-%m-%d").date()
    days = (td - fd).days + 1
    leave = Leave(employee_id=data.employee_id, leave_type=data.leave_type,
                  from_date=fd, to_date=td, days=days, reason=data.reason)
    db.add(leave); db.commit(); db.refresh(leave)
    return {"id": leave.id, "days": leave.days, "status": leave.status}

@router.put("/leaves/{leave_id}/approve")
async def approve_leave(leave_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from datetime import datetime
    l = db.query(Leave).filter(Leave.id == leave_id).first()
    if not l: raise HTTPException(404, "Not found")
    l.status = "approved"; l.approved_by = current_user.id; l.approved_at = datetime.utcnow()
    db.commit(); return {"message": "Approved"}

@router.put("/leaves/{leave_id}/reject")
async def reject_leave(leave_id: int, reason: str = "", db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    l = db.query(Leave).filter(Leave.id == leave_id).first()
    if not l: raise HTTPException(404, "Not found")
    l.status = "rejected"; l.rejection_reason = reason; db.commit()
    return {"message": "Rejected"}

# ─── Payroll ─────────────────────────────────────────────────────────────────

@router.get("/payroll")
async def list_payroll(skip: int = 0, limit: int = 50, db: Session = Depends(get_db),
                        current_user: User = Depends(get_current_user)):
    q = db.query(Payroll)
    total = q.count()
    items = q.order_by(Payroll.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [{"id": p.id, "employee_id": p.employee_id,
            "payroll_period": p.payroll_period, "gross_salary": float(p.gross_salary or 0),
            "net_salary": float(p.net_salary or 0), "status": p.status} for p in items]}

@router.post("/payroll/run")
async def run_payroll(month: int, year: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    employees = db.query(Employee).filter(Employee.is_active == True, Employee.status == "active").all()
    created = []
    for emp in employees:
        period = f"{year}-{month:02d}"
        existing = db.query(Payroll).filter(Payroll.employee_id == emp.id, Payroll.payroll_period == period).first()
        if existing: continue
        gross = float(emp.basic_salary or 0) + float(emp.hra or 0) + float(emp.other_allowances or 0)
        pf_emp = float(emp.basic_salary or 0) * 0.12
        esi_emp = gross * 0.0075 if gross <= 21000 else 0
        professional_tax = 200
        total_ded = pf_emp + esi_emp + professional_tax
        net = gross - total_ded
        p = Payroll(employee_id=emp.id, payroll_period=period, month=month, year=year,
                     basic_salary=emp.basic_salary, hra=emp.hra, other_allowances=emp.other_allowances,
                     gross_salary=gross, pf_employee=pf_emp, esic_employee=esi_emp,
                     professional_tax=professional_tax, total_deductions=total_ded, net_salary=net)
        db.add(p); created.append(emp.id)
    db.commit()
    return {"message": f"Payroll generated for {len(created)} employees", "period": f"{year}-{month:02d}"}

@router.put("/payroll/{payroll_id}")
async def update_payroll(payroll_id: int, data: dict, db: Session = Depends(get_db),
                          current_user: User = Depends(get_current_user)):
    p = db.query(Payroll).filter(Payroll.id == payroll_id).first()
    if not p: raise HTTPException(404, "Not found")
    for k, v in data.items():
        if hasattr(p, k): setattr(p, k, v)
    db.commit()
    return {"id": p.id, "status": p.status, "payroll_period": p.payroll_period,
            "gross_salary": float(p.gross_salary or 0), "net_salary": float(p.net_salary or 0)}

@router.get("/payroll/export")
async def export_payroll(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    items = db.query(Payroll).order_by(Payroll.payroll_period.desc()).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Payroll"
    headers = ["Employee ID", "Period", "Gross Salary", "PF", "ESI", "Professional Tax", "Total Deductions", "Net Salary", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6366F1")
    for row, p in enumerate(items, 2):
        for col, val in enumerate([p.employee_id, p.payroll_period, float(p.gross_salary or 0),
                                    float(p.pf_employee or 0), float(p.esic_employee or 0),
                                    float(p.professional_tax or 0), float(p.total_deductions or 0),
                                    float(p.net_salary or 0), p.status], 1):
            ws.cell(row=row, column=col, value=val)
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=payroll.xlsx"})
