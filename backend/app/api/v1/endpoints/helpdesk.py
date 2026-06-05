from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from pydantic import BaseModel
import io, openpyxl
from openpyxl.styles import Font, PatternFill

from app.db.base import get_db
from app.models.helpdesk import Ticket, TicketComment, SLA, KnowledgeBase
from app.models.user import User
from app.api.v1.endpoints.auth import get_current_user

router = APIRouter()

class TicketCreate(BaseModel):
    subject: str
    description: Optional[str] = None
    ticket_type: str = "incident"
    category: Optional[str] = None
    priority: str = "medium"
    customer_id: Optional[int] = None
    contact_name: Optional[str] = None
    contact_email: Optional[str] = None
    contact_phone: Optional[str] = None

def ticket_dict(t: Ticket):
    return {"id": t.id, "ticket_number": t.ticket_number, "subject": t.subject,
            "ticket_type": t.ticket_type, "category": t.category, "priority": t.priority,
            "status": t.status, "customer_id": t.customer_id, "contact_name": t.contact_name,
            "contact_email": t.contact_email, "assigned_to": t.assigned_to,
            "sla_breached": t.sla_breached, "is_active": t.is_active,
            "created_at": t.created_at.isoformat() if t.created_at else None,
            "resolved_at": t.resolved_at.isoformat() if t.resolved_at else None}

@router.get("/tickets")
async def list_tickets(skip: int = 0, limit: int = 50, status: Optional[str] = None,
                        priority: Optional[str] = None, assigned_to: Optional[int] = None,
                        db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    q = db.query(Ticket).filter(Ticket.is_active == True)
    if status: q = q.filter(Ticket.status == status)
    if priority: q = q.filter(Ticket.priority == priority)
    if assigned_to: q = q.filter(Ticket.assigned_to == assigned_to)
    total = q.count()
    items = q.order_by(Ticket.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [ticket_dict(t) for t in items]}

@router.post("/tickets")
async def create_ticket(data: TicketCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    count = db.query(func.count(Ticket.id)).scalar()
    t = Ticket(ticket_number=f"TKT-{(count or 0) + 1:05d}", **data.dict())
    db.add(t); db.commit(); db.refresh(t)
    return ticket_dict(t)

@router.get("/tickets/stats")
async def ticket_stats(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    total = db.query(func.count(Ticket.id)).filter(Ticket.is_active == True).scalar()
    by_status = dict(db.query(Ticket.status, func.count(Ticket.id)).filter(Ticket.is_active == True).group_by(Ticket.status).all())
    by_priority = dict(db.query(Ticket.priority, func.count(Ticket.id)).filter(Ticket.is_active == True).group_by(Ticket.priority).all())
    breached = db.query(func.count(Ticket.id)).filter(Ticket.sla_breached == True, Ticket.is_active == True).scalar()
    return {"total": total, "by_status": by_status, "by_priority": by_priority, "sla_breached": breached}

@router.get("/tickets/export")
async def export_tickets(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    items = db.query(Ticket).filter(Ticket.is_active == True).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Tickets"
    headers = ["Ticket #", "Subject", "Type", "Category", "Priority", "Status", "Contact", "Created"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6366F1")
    for row, t in enumerate(items, 2):
        for col, val in enumerate([t.ticket_number, t.subject, t.ticket_type, t.category,
                                    t.priority, t.status, t.contact_name,
                                    t.created_at.strftime("%Y-%m-%d") if t.created_at else ""], 1):
            ws.cell(row=row, column=col, value=val)
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=tickets.xlsx"})

@router.get("/tickets/{ticket_id}")
async def get_ticket(ticket_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    t = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not t: raise HTTPException(404, "Not found")
    d = ticket_dict(t)
    d["comments"] = [{"id": c.id, "comment": c.comment, "is_internal": c.is_internal,
                      "created_by": c.created_by, "created_at": c.created_at.isoformat()} for c in t.comments]
    return d

@router.put("/tickets/{ticket_id}")
async def update_ticket(ticket_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from datetime import datetime
    t = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not t: raise HTTPException(404, "Not found")
    for k, v in data.items():
        if hasattr(t, k): setattr(t, k, v)
    if data.get("status") in ["resolved", "closed"] and not t.resolved_at:
        t.resolved_at = datetime.utcnow()
    db.commit(); return ticket_dict(t)

@router.post("/tickets/{ticket_id}/comments")
async def add_comment(ticket_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    c = TicketComment(ticket_id=ticket_id, comment=data.get("comment"),
                      is_internal=data.get("is_internal", False), created_by=current_user.id)
    db.add(c); db.commit(); db.refresh(c)
    return {"id": c.id}

# ─── Knowledge Base ───────────────────────────────────────────────────────────

@router.get("/knowledge")
async def list_kb(skip: int = 0, limit: int = 50, search: Optional[str] = None,
                   category: Optional[str] = None, db: Session = Depends(get_db),
                   current_user: User = Depends(get_current_user)):
    q = db.query(KnowledgeBase).filter(KnowledgeBase.is_active == True, KnowledgeBase.status == "published")
    if search: q = q.filter(KnowledgeBase.title.ilike(f"%{search}%"))
    if category: q = q.filter(KnowledgeBase.category == category)
    total = q.count()
    items = q.order_by(KnowledgeBase.views.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [{"id": k.id, "title": k.title, "category": k.category,
            "views": k.views, "helpful_count": k.helpful_count, "tags": k.tags} for k in items]}

@router.post("/knowledge")
async def create_kb(data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    kb = KnowledgeBase(title=data.get("title"), content=data.get("content"),
                        category=data.get("category"), tags=data.get("tags", []),
                        created_by=current_user.id)
    db.add(kb); db.commit(); db.refresh(kb)
    return {"id": kb.id}

@router.get("/knowledge/{kb_id}")
async def get_kb(kb_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    kb = db.query(KnowledgeBase).filter(KnowledgeBase.id == kb_id).first()
    if not kb: raise HTTPException(404, "Not found")
    kb.views += 1; db.commit()
    return {"id": kb.id, "title": kb.title, "content": kb.content, "category": kb.category,
            "tags": kb.tags, "views": kb.views, "helpful_count": kb.helpful_count}
